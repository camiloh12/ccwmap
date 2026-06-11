"""GSA FRPP (Federal Real Property Profile) -> Candidate stream.

Source: GSA FRPP public dataset. Federal owned/leased real property.
License: public domain (US Gov). Category: FEDERAL_PROPERTY (18 USC 930 via the
US FEDERAL_PROPERTY state-law cell, source_filter includes gsa).

Format (verified 2026-06 against the FY24 public release): a single-sheet
**.xlsx** (~145 MB, ~308k rows, 117 columns), streamed read-only with openpyxl —
NOT a CSV. Column headers and a few semantic facts that drove this module:
  - State is carried as `State Name` (UPPERCASE full name, e.g. "TEXAS"); the
    `State Code` column is FIPS-numeric, so we map the requested USPS codes to
    full names before filtering.
  - There is no per-asset name column; `Installation Name` is the closest, so
    many buildings on one base share a name (acceptable; dedup is by external id).
  - `Real Property Type` is Building / Structure / Land. Only **Building** is a
    "federal facility" in the 18 USC 930 sense, so Structures and Land are
    dropped (see _INCLUDED_TYPES).
  - `Latitude`/`Longitude` are populated on ~97% of rows, so the Census geocoder
    is a rare fallback for the address-only remainder.
"""

from __future__ import annotations

import re
from collections import Counter
from collections.abc import Iterator
from pathlib import Path
from typing import ClassVar, Protocol

import httpx
import openpyxl

from importer.candidate import Candidate, CoordQuality
from importer.geo.census_geocode import AddressRecord
from importer.geo.states import StateLocator
from importer.restriction_tag import RestrictionTag
from importer.sources.base import Source

_COL_ID = "Real Property Unique Identifier"
_COL_TYPE = "Real Property Type"
_COL_STREET = "Street Address"
_COL_CITY = "City Name"
_COL_STATE = "State Name"
_COL_ZIP = "Zip Code"
_COL_LAT = "Latitude"
_COL_LNG = "Longitude"
_COL_NAME = "Installation Name"
_COL_USE = "Real Property Use"

# Only the columns this module reads — keeps the per-row dict small over ~308k rows.
_NEEDED_COLS = (
    _COL_ID, _COL_TYPE, _COL_STREET, _COL_CITY, _COL_STATE, _COL_ZIP,
    _COL_LAT, _COL_LNG, _COL_NAME, _COL_USE,
)

# 18 USC 930 attaches to federal *facilities* (buildings), not Structures or Land.
_INCLUDED_TYPES = {"building"}

# FRPP's only name column, Installation Name, is occasionally just a city
# ("TAMPA, FL") or a street address ("9450 Koger Boulevard"). Detect those and
# compose a clearer label from Real Property Use + City instead.
_CITY_LIKE = re.compile(r"^[A-Za-z][A-Za-z .'\-]+,\s*[A-Za-z]{2}$")


def _display_name(installation_name: str, use: str, city: str, usps: str) -> str:
    nm = installation_name.strip()
    is_degenerate = bool(_CITY_LIKE.match(nm)) or nm[:1].isdigit()
    if not is_degenerate:
        return nm
    use_label = use.strip()
    if not use_label or use_label.lower() == "all other":
        use_label = "Federal property"
    city_label = city.strip().title()
    return f"{use_label} — {city_label}, {usps}" if city_label else f"{use_label}, {usps}"

# FRPP's State Name is the UPPERCASE full name; we filter on USPS codes.
_USPS_TO_STATE_NAME: dict[str, str] = {
    "AL": "ALABAMA", "AK": "ALASKA", "AZ": "ARIZONA", "AR": "ARKANSAS",
    "CA": "CALIFORNIA", "CO": "COLORADO", "CT": "CONNECTICUT", "DE": "DELAWARE",
    "DC": "DISTRICT OF COLUMBIA", "FL": "FLORIDA", "GA": "GEORGIA", "HI": "HAWAII",
    "ID": "IDAHO", "IL": "ILLINOIS", "IN": "INDIANA", "IA": "IOWA",
    "KS": "KANSAS", "KY": "KENTUCKY", "LA": "LOUISIANA", "ME": "MAINE",
    "MD": "MARYLAND", "MA": "MASSACHUSETTS", "MI": "MICHIGAN", "MN": "MINNESOTA",
    "MS": "MISSISSIPPI", "MO": "MISSOURI", "MT": "MONTANA", "NE": "NEBRASKA",
    "NV": "NEVADA", "NH": "NEW HAMPSHIRE", "NJ": "NEW JERSEY", "NM": "NEW MEXICO",
    "NY": "NEW YORK", "NC": "NORTH CAROLINA", "ND": "NORTH DAKOTA", "OH": "OHIO",
    "OK": "OKLAHOMA", "OR": "OREGON", "PA": "PENNSYLVANIA", "RI": "RHODE ISLAND",
    "SC": "SOUTH CAROLINA", "SD": "SOUTH DAKOTA", "TN": "TENNESSEE", "TX": "TEXAS",
    "UT": "UTAH", "VT": "VERMONT", "VA": "VIRGINIA", "WA": "WASHINGTON",
    "WV": "WEST VIRGINIA", "WI": "WISCONSIN", "WY": "WYOMING",
}
_STATE_NAME_TO_USPS: dict[str, str] = {v: k for k, v in _USPS_TO_STATE_NAME.items()}


class _Geocoder(Protocol):
    def geocode(self, records: list[AddressRecord]) -> dict[str, tuple[float, float]]: ...


class GsaSource(Source):
    SOURCE_NAME: ClassVar[str] = "gsa"

    def __init__(
        self,
        *,
        cache_path: Path,
        dataset_version: str,
        geocoder: _Geocoder,
        state_locator: StateLocator,
        url: str = "",
    ) -> None:
        self._cache_path = cache_path
        self._version = dataset_version
        self._geocoder = geocoder
        self._locator = state_locator
        self._url = url
        self.last_skip_counts: Counter[str] = Counter()

    def fetch(self, *, refetch: bool = False) -> None:
        if self._cache_path.exists() and not refetch:
            return
        if not self._url:
            raise RuntimeError("gsa url not configured; set sources.gsa.url in config.yaml")
        self._cache_path.parent.mkdir(parents=True, exist_ok=True)
        with httpx.Client(timeout=300.0, follow_redirects=True) as client:
            r = client.get(self._url)
            r.raise_for_status()
            self._cache_path.write_bytes(r.content)

    def _read_rows(self) -> Iterator[dict[str, str]]:
        """Stream the .xlsx, yielding one dict per data row for the needed columns.

        All values are stringified (openpyxl returns floats/ints for numeric
        cells) so downstream `.strip()`/`float()` logic is uniform.
        """
        wb = openpyxl.load_workbook(self._cache_path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            header = ["" if h is None else str(h).strip() for h in next(it)]
            col_idx = {name: header.index(name) for name in _NEEDED_COLS if name in header}
            missing = [name for name in _NEEDED_COLS if name not in col_idx]
            if missing:
                raise RuntimeError(f"GSA FRPP xlsx missing expected columns: {missing}")
            for raw in it:
                yield {
                    name: ("" if raw[i] is None else str(raw[i]).strip())
                    for name, i in col_idx.items()
                }
        finally:
            wb.close()

    def iter_candidates(self, state_filter: set[str] | None) -> Iterator[Candidate]:
        self.last_skip_counts = Counter()
        wanted_names = (
            {_USPS_TO_STATE_NAME[c] for c in state_filter if c in _USPS_TO_STATE_NAME}
            if state_filter is not None
            else None
        )

        kept: list[dict] = []
        to_geocode: list[AddressRecord] = []
        for row in self._read_rows():
            state_name = (row.get(_COL_STATE) or "").strip().upper()
            usps = _STATE_NAME_TO_USPS.get(state_name)
            if wanted_names is not None and state_name not in wanted_names:
                self.last_skip_counts["filtered_out"] += 1
                continue
            if usps is None:
                # Territories / blanks have no USPS code and no state-law cell.
                self.last_skip_counts["unmapped_state"] += 1
                continue
            if (row.get(_COL_TYPE) or "").strip().lower() not in _INCLUDED_TYPES:
                self.last_skip_counts["not_federal_facility"] += 1
                continue
            eid = (row.get(_COL_ID) or "").strip()
            if not eid:
                self.last_skip_counts["missing_external_id"] += 1
                continue
            name = (row.get(_COL_NAME) or "").strip()
            if not name:
                self.last_skip_counts["missing_name"] += 1
                continue
            street = (row.get(_COL_STREET) or "").strip()
            lat, lng = self._existing_coords(row)
            if lat is None and not street:
                self.last_skip_counts["missing_address"] += 1
                continue
            row["_usps"] = usps
            row["_eid"] = eid
            row["_name"] = name
            row["_lat"] = lat
            row["_lng"] = lng
            kept.append(row)
            if lat is None:
                to_geocode.append(AddressRecord(
                    id=eid, street=street,
                    city=(row.get(_COL_CITY) or "").strip(),
                    state=usps, zip=(row.get(_COL_ZIP) or "").strip(),
                ))

        geocoded = self._geocoder.geocode(to_geocode) if to_geocode else {}

        for row in kept:
            lat, lng = row["_lat"], row["_lng"]
            quality = CoordQuality.PRECISE
            if lat is None:
                hit = geocoded.get(row["_eid"])
                if hit is None:
                    self.last_skip_counts["geocode_miss"] += 1
                    continue
                lat, lng = hit
                quality = CoordQuality.ADDRESS_CENTROID
            # Defense-in-depth: the FRPP State Name field and the Latitude/
            # Longitude columns can disagree (corrupt coord, (0,0) null-island).
            # Trust the geometry — if the coordinate doesn't fall inside the
            # claimed state, drop the row rather than plot a pin in the wrong place.
            if self._locator.state_for(lat, lng) != row["_usps"]:
                self.last_skip_counts["coord_state_mismatch"] += 1
                continue
            yield Candidate(
                source=self.SOURCE_NAME,
                source_external_id=row["_eid"],
                source_dataset_version=self._version,
                name=_display_name(
                    row["_name"], row.get(_COL_USE, ""), row.get(_COL_CITY, ""), row["_usps"]
                ),
                latitude=lat,
                longitude=lng,
                coord_quality=quality,
                category=RestrictionTag.FEDERAL_PROPERTY,
                state=row["_usps"],
                extra={},
            )

    @staticmethod
    def _existing_coords(row: dict) -> tuple[float | None, float | None]:
        lat_s = (row.get(_COL_LAT) or "").strip()
        lng_s = (row.get(_COL_LNG) or "").strip()
        if not lat_s or not lng_s:
            return None, None
        try:
            return float(lat_s), float(lng_s)
        except ValueError:
            return None, None
